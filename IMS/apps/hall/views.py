import pandas as pd
from django.shortcuts import render,redirect
from .models import Room
from django.db.models import Sum
from django.db.models import Sum, Count
from staff.logic import allocate_sessions
from django.http import HttpResponse
from openpyxl import Workbook
from apps.staff.models import Staff
from django.utils import timezone
import datetime
from collections import defaultdict
from ..exam_dates.models import ExamDate
from ..invigilation_schedule.models import InvigilationSchedule
import random
import openpyxl
from django.db.models import Q, F, Count, Case, When
from django.contrib import messages
import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from django.db import transaction
from django.views.decorators.http import require_POST
from django.http import JsonResponse

from .slot_optimizer import reduce_unassigned_slots
from .reduce_unassigned_slots import reduce_unassigned_slots

def is_assignment_allowed(current_s1, current_s2, new_session):
    """
    Checks assignment rules in reverse order:
    1. Allowed (best) → 2. Avoidable → 3. Blocked (worst)
    """
    difference = (current_s1 - current_s2)
    
    # Calculate projected difference
    projected_diff = difference + 1 if new_session == 1 else difference - 1

    # 1. First check for ideal cases (0 or ±1 difference)
    if abs(projected_diff) <= 1:
        return True, "Allowed (imbalance ≤ 1)"  # ✅ Best case

    # 2. Then check avoidable cases (±2 difference)
    elif abs(projected_diff) == 2:
        return True, "Allowed but avoid (imbalance 2)"  # ⚠️ Good but not ideal

    # 3. Finally handle blocked cases (≥±3 difference)
    else:
        return False, "Blocked (imbalance ≥ 3)"  # ❌ Worst case

def hall_management(request):
    print("\n===== NEW REQUEST =====")  # Debug
    print(f"Request method: {request.method}")  # Debug
    
    # Initialize context with all required data
    context = {
        'rooms': Room.objects.all(),
        'blocks': Room.objects.values_list('block', flat=True).distinct(),
        'dept_types': Room.objects.values_list('dept_category', flat=True).distinct(),
        'departments': Room.objects.values_list('dept_name', flat=True).distinct(),
    }

    if request.method == "POST":
        print("\nPOST request received")  # Debug
        print("Request.FILES contents:", request.FILES)  # Debug
        print("Request.POST contents:", request.POST)  # Debug
        
        if 'file' not in request.FILES:
            print("\nERROR: No file found in request.FILES")  # Debug
            context['error'] = "No file was uploaded"
            return render(request, 'hall/hall_management.html', context)
            
        excel_file = request.FILES['file']
        print(f"\nFile detected: {excel_file.name}")  # Debug
        print(f"File size: {excel_file.size} bytes")  # Debug
        print(f"Content type: {excel_file.content_type}")  # Debug
        
        try:
            print("\nAttempting to read Excel file...")  # Debug
            df = pd.read_excel(excel_file)
            total = len(df)
            # Delete existing rooms BEFORE processing new data
            if Room.objects.exists():  # Safety check
                deleted_count = Room.objects.all().delete()[0]
            print("SUCCESS: File read successfully")  # Debug
            print("Columns in file:", df.columns.tolist())  # Debug
            print("First row sample:", df.iloc[0].to_dict() if not df.empty else "Empty DataFrame")  # Debug

            processed = 0
            for _, row in df.iterrows():
                try:
                    hall_no = str(row['hall_no']).strip()
                    if not hall_no:
                        continue
                    
                    # Calculate staff_required based on strength
                    strength = int(row.get('strength', 0))
                    staff_required = 1 if strength < 55 else 2
                    # Get days value from Excel (default to 1 if not provided)
                    days = int(row.get('days', 1))
                    
                    # Calculate required_session
                    sessions = 2  # Assuming 2 sessions per day as per your requirement
                    required_session = staff_required * sessions * days
                    
                    
                    Room.objects.create(
                        hall_no=hall_no,
                        dept_category=str(row.get('dept_category', '')),
                        dept_name=str(row.get('dept_name', '')),
                        strength=strength,
                        days=int(row.get('days', 0)),
                        staff_required=staff_required,
                        required_session=required_session,
                        block="-",
                        staff_allotted="Not Allotted",
                        benches=0,
                    )
                    processed += 1
                except Exception as e:
                    print(f"Row processing error: {e}")  # Debug
                    continue

            context['message'] = f"Successfully processed {processed} / {total} records"
                # Refresh data
            context.update({
                'rooms': Room.objects.all(),
                'blocks': Room.objects.values_list('block', flat=True).distinct(),
                'dept_types': Room.objects.values_list('dept_category', flat=True).distinct(),
                'departments': Room.objects.values_list('dept_name', flat=True).distinct(),
            })
            
        except Exception as e:
            print(f"\nERROR processing file: {str(e)}")  # Debug
            context['error'] = f"Error processing file: {str(e)}"

    return render(request, 'hall/hall_management.html', context)

def hall_list(request):
    rooms = Room.objects.all()
    return render(request, 'hall/hall_list.html', {'halls': rooms})

def generate_schedule(request):
    


    CATEGORIES = ['AIDED', 'SFM', 'SFW']
    
    # # Get room data
    room_data = Room.objects.values('dept_category').annotate(
        total_sessions=Sum('required_session')
    ).order_by('dept_category')
    
    # # Get staff data
    staff_data = Staff.objects.values('dept_category').annotate(
        total_staff=Count('id')
    ).order_by('dept_category')
    
    room_dict = {item['dept_category'].upper(): item for item in room_data}
    staff_dict = {item['dept_category'].upper(): item for item in staff_data}
    
    # Prepare schedule data for display
    schedule_data = []
    for category in CATEGORIES:
        sessions = room_dict.get(category, {}).get('total_sessions', 0)
        staff_count = staff_dict.get(category, {}).get('total_staff', 0)
        
        schedule_data.append({
            'category': category,
            'total_sessions': sessions,
            'staff_count': staff_count
        })

    required_session = {
        item['dept_category']: item['total_sessions']
        for item in room_data
    }
    
    # # Handle allocation if requested
    allocation_results = None
    if request.method == 'POST' and 'allocate_sessions' in request.POST:
        allocation_results = allocate_sessions(required_session)

    # Calculate total sessions required by category
    categories = Room.objects.values('dept_category').annotate(
        total_sessions=Sum('required_session')
    ).order_by('dept_category')

    # # Get staff counts with case-insensitive matching  
    staff_counts = {
        'Aided': Staff.objects.filter(dept_category__iexact='aided').count(),
        'SFM': Staff.objects.filter(dept_category__iexact='sfm').count(),
        'SFW': Staff.objects.filter(dept_category__iexact='sfw').count()
    }
    
    # Format the data for display with staff counts
    schedule_data = []
    for item in categories:
            dept = item['dept_category']
            # Normalize department name for lookup
            lookup_dept = 'Aided' if dept.lower() == 'aided' else dept.upper()
            schedule_data.append({
                'category': dept,
                'total_sessions': item['total_sessions'],
                'staff_count': staff_counts.get(lookup_dept, 0),
                'allocation_status': allocation_results[dept]['status'] if allocation_results else None
            })

    context = {
        'schedule_data': schedule_data,
        'categories': CATEGORIES,
        'allocation_results': allocation_results
    }

    return render(request, 'hall/generate_schedule.html', context)


def export_schedule_excel(request):
    # Create a workbook and add worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Schedule"
    
    # Add column headers
    headers = [
        "Staff ID",
        "Name", 
        "Department",
        "Department Category",
        "Sessions",
        "Join Date",
        "Fixed session",
    ]
    ws.append(headers)
    
    # Get all staff data with correct comma separation
    staff_data = Staff.objects.all().values_list(
        'staff_id',      # Staff ID
        'name',          # Name
        'dept_name',     # Department
        'dept_category', # Department Category
        'session',       # Sessions
        'date_of_joining', # Join Date
        'fixed_session',
    )
    
    # Add data rows
    for staff in staff_data:
        ws.append(staff)
    
    # Format the date fields
    for row in ws.iter_rows(min_row=2):
        if len(row) > 5 and isinstance(row[5].value, (datetime.date, datetime.datetime)):  # Date of Joining column (6th column)
            row[5].number_format = 'YYYY-MM-DD'
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"Staff_Session_Calculation.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def generate_session(request):
    InvigilationSchedule.objects.all().delete()
    if request.method == "POST":
        print("this is the route")
        halls = Room.objects.all().values() 
        dates=ExamDate.objects.all()
        print("dates",dates)

        # Assuming you already have dates and halls loaded

        for date in dates:
            for hall in halls:
                total = hall['staff_required'] * 2
                for i in range(total):
                    session_number = "1" if i < total // 2 else "2"
                    # print("gsfd",date.date)

                    InvigilationSchedule.objects.create(
                        date=date.date,
                        session=session_number,
                        hall_no=hall['hall_no'],
                        hall_department=hall['dept_name'],
                        hall_dept_category=hall['dept_category'],
                        #set later
                        dept_category="None",
                        staff_id=None,
                        name=None,
                        designation=None,
                        staff_category=None,
                        double_session=False
                    )
                    # print("gsfd",date.date)
        return redirect('generate_schedule')

#NEW 

def assign_staff(request):
    if request.method != 'POST':
        return redirect('generate_schedule')

    exam_dates = list(ExamDate.objects.all().order_by('date'))
    total_dates = len(exam_dates)

    # eligible_staff = Staff.objects.filter(is_active=True, session__gt=0).order_by('-session', 'random')
    # eligible_staff = Staff.objects.filter(is_active=True, session__gt=0).order_by('-session')
    from django.db.models import Count

    eligible_staff = (
        Staff.objects
        .filter(is_active=True, session__gt=0)
        .annotate(dept_size=Count('dept_name'))
        .order_by('-session', '-dept_size')
    )


    # 🔧 Helper: Get matched slots with progressive fallback
    def get_matched_slots(staff):
        # 🎯 Primary filter: same category, different department
        slots = InvigilationSchedule.objects.filter(
            staff_id__isnull=True,
            hall_dept_category=staff.dept_category
        ).exclude(
            hall_department=staff.dept_name
        )
        if slots.exists():
            return slots

        # 🔁 Fallback: same category, same department
        slots = InvigilationSchedule.objects.filter(
            staff_id__isnull=True,
            hall_dept_category=staff.dept_category,
            hall_department=staff.dept_name
        )
        if slots.exists():
            print(f"🔁 Relaxed department constraint for {staff.name} ({staff.staff_id})")
            return slots

        # 🛠️ Final fallback: ignore category and department
        slots = InvigilationSchedule.objects.filter(
            staff_id__isnull=True
        )
        if slots.exists():
            print(f"🛠️ Final fallback used for {staff.name} ({staff.staff_id})")
            return slots

        return InvigilationSchedule.objects.none()

    # 🔧 Helper: Get valid slot for a specific date
    def get_valid_slot(staff, date, matched_slots):
        assigned_sessions = InvigilationSchedule.objects.filter(
            staff_id=staff.staff_id,
            date=date.date
        ).values_list('session', flat=True)

        opposite_session = None
        if '1' in assigned_sessions:
            opposite_session = '2'
        elif '2' in assigned_sessions:
            opposite_session = '1'

        if opposite_session:
            hall_nos_on_date = InvigilationSchedule.objects.filter(
                staff_id=staff.staff_id,
                date=date.date
            ).values_list('hall_no', flat=True)

            slot = matched_slots.filter(
                date=date.date,
                session=opposite_session,
                hall_no__in=hall_nos_on_date
            ).first()
            if slot:
                return slot

            slot = matched_slots.filter(
                date=date.date,
                session=opposite_session
            ).first()
            if slot:
                return slot

            slot = matched_slots.filter(
                session=opposite_session
            ).first()
            if slot:
                return slot

        slot = matched_slots.filter(
            date=date.date
        ).exclude(
            session__in=assigned_sessions
        ).first()
        return slot

    # 🔧 Helper: Overflow slot logic
    def get_overflow_slot(staff, matched_slots):
        assigned = InvigilationSchedule.objects.filter(staff_id=staff.staff_id)

        for record in assigned:
            opp_session = '2' if record.session == '1' else '1'
            slot = matched_slots.filter(
                date=record.date,
                hall_no=record.hall_no,
                session=opp_session
            ).first()
            if slot:
                return slot

        for record in assigned:
            opp_session = '2' if record.session == '1' else '1'
            slot = matched_slots.filter(
                hall_no=record.hall_no,
                session=opp_session
            ).exclude(date=record.date).first()
            if slot:
                return slot

        for record in assigned:
            opp_session = '2' if record.session == '1' else '1'
            slot = matched_slots.filter(session=opp_session).first()
            if slot:
                return slot

        return None
    
    assigned_slots_set = set()  # To track (staff_id, date, session, hall_no) tuples assigned during this run

    # 🚀 Begin assignment loop
    for staff in eligible_staff:
        current_assignments = InvigilationSchedule.objects.filter(staff_id=staff.staff_id).count()
        remaining_assignments = staff.session - current_assignments
        if remaining_assignments <= 0:
            continue

        matched_slots = get_matched_slots(staff)
        if matched_slots.count() == 0:
            continue

        if remaining_assignments > total_dates:
            for date in exam_dates:
                slot = get_valid_slot(staff, date, matched_slots)
                if slot:
                    # Prevent duplicate assignment to same hall, session, and date for the staff
                    key = (staff.staff_id, slot.date, slot.session, slot.hall_no)
                    if key in assigned_slots_set:
                        continue  # Skip duplicate assignment
  # skip duplicate slot

                    slot.staff_id = staff.staff_id
                    slot.name = staff.name
                    slot.designation = str(staff.designation)
                    slot.staff_category = staff.staff_category
                    slot.dept_category = staff.dept_category
                    slot.dept_name = staff.dept_name  # ADDED: Include dept_name
                    slot.save()
                    assigned_slots_set.add(key)
                    matched_slots = matched_slots.exclude(pk=slot.pk)

            overflow_count = remaining_assignments - total_dates
            assigned = 0
            while assigned < overflow_count:
                slot = get_overflow_slot(staff, matched_slots)
                if not slot:
                    break

                key = (staff.staff_id, slot.date, slot.session, slot.hall_no)

                # Check in-memory duplicate
                if key in assigned_slots_set:
                    matched_slots = matched_slots.exclude(pk=slot.pk)
                    continue

                # Check in database duplicate
                if InvigilationSchedule.objects.filter(
                    staff_id=staff.staff_id,
                    date=slot.date,
                    session=slot.session,
                    hall_no=slot.hall_no
                ).exists():
                    matched_slots = matched_slots.exclude(pk=slot.pk)
                    continue
                slot.staff_id = staff.staff_id
                slot.name = staff.name
                slot.designation = str(staff.designation)
                slot.staff_category = staff.staff_category
                slot.dept_category = staff.dept_category
                slot.dept_name = staff.dept_name  # ADDED: Include dept_name
                slot.save()
                
                matched_slots = matched_slots.exclude(pk=slot.pk)
                assigned += 1

        elif remaining_assignments == total_dates:
            session_tracker = {'1': 0, '2': 0}
            assigned_departments = set()

            for date in exam_dates:
                preferred_session = '1' if session_tracker['1'] <= session_tracker['2'] else '2'

                # Filter slots for this date, preferred session, and unassigned department
                preferred_slots = matched_slots.filter(
                    date=date.date,
                    session=preferred_session
                ).exclude(
                    hall_department=staff.dept_name
                ).exclude(
                    hall_department__in=assigned_departments
                )

                slot = preferred_slots.first()

                # Fallback 1: allow repeated departments but still exclude own
                if not slot:
                    fallback_slots = matched_slots.filter(
                        date=date.date,
                        session=preferred_session
                    ).exclude(
                        hall_department=staff.dept_name
                    )
                    slot = fallback_slots.first()

                # Fallback 2: any slot on that date
                if not slot:
                    slot = matched_slots.filter(date=date.date).exclude(
                        hall_department=staff.dept_name
                    ).first()

                if slot:
                    # Prevent duplicate assignment to same hall, session, and date for the staff
                    key = (staff.staff_id, slot.date, slot.session, slot.hall_no)
                    if key in assigned_slots_set:
                        continue  # skip duplicate slot

                    slot.staff_id = staff.staff_id
                    slot.name = staff.name
                    slot.designation = str(staff.designation)
                    slot.staff_category = staff.staff_category
                    slot.dept_category = staff.dept_category
                    slot.dept_name = staff.dept_name  # ADDED: Include dept_name
                    slot.save()
                    assigned_slots_set.add(key)
                    matched_slots = matched_slots.exclude(pk=slot.pk)


                    session_tracker[slot.session] += 1
                    assigned_departments.add(slot.hall_department)

        elif remaining_assignments < total_dates:
            selected_dates = random.sample(exam_dates, remaining_assignments)
            # selected_dates = exam_dates[:remaining_assignments]
            session_tracker = {'1': 0, '2': 0}
            assigned_departments = set()

            for date in selected_dates:
                preferred_session = '1' if session_tracker['1'] <= session_tracker['2'] else '2'

                # Filter slots for this date, preferred session, and unassigned department
                preferred_slots = matched_slots.filter(
                    date=date.date,
                    session=preferred_session
                ).exclude(
                    hall_department=staff.dept_name
                ).exclude(
                    hall_department__in=assigned_departments
                )

                slot = preferred_slots.first()

                # Fallback 1: allow repeated departments but still exclude own
                if not slot:
                    fallback_slots = matched_slots.filter(
                        date=date.date,
                        session=preferred_session
                    ).exclude(
                        hall_department=staff.dept_name
                    )
                    slot = fallback_slots.first()

                # Fallback 2: any slot on that date
                if not slot:
                    slot = matched_slots.filter(date=date.date).exclude(
                        hall_department=staff.dept_name
                    ).first()

                if slot:
                    key = (staff.staff_id, slot.date, slot.session, slot.hall_no)
                    if key in assigned_slots_set:
                        continue  # Skip duplicate assignment
                    # Prevent duplicate assignment to same hall, session, and date for the staff
                    duplicate_assignment = InvigilationSchedule.objects.filter(
                    staff_id=staff.staff_id,
                    date=slot.date,
                    session=slot.session,
                    hall_no=slot.hall_no
                ).exists()
                    if duplicate_assignment:
                        continue  # skip duplicate slot

                    slot.staff_id = staff.staff_id
                    slot.name = staff.name
                    slot.designation = str(staff.designation)
                    slot.staff_category = staff.staff_category
                    slot.dept_category = staff.dept_category
                    slot.dept_name = staff.dept_name  # ADDED: Include dept_name
                    slot.save()
                    assigned_slots_set.add(key)
                    matched_slots = matched_slots.exclude(pk=slot.pk)

                    session_tracker[slot.session] += 1
                    assigned_departments.add(slot.hall_department)

    # 📊 Final report
    total_slots = InvigilationSchedule.objects.count()
    unassigned = InvigilationSchedule.objects.filter(staff_id__isnull=True).count()
    print(f"✅ Staff assignment complete. 📊 Coverage: {total_slots - unassigned}/{total_slots} slots assigned")
    if unassigned > 0:
        # Get all department categories with unassigned slots
        dept_categories_with_unassigned = InvigilationSchedule.objects.filter(
            staff_id__isnull=True
        ).values_list('hall_dept_category', flat=True).distinct()
        
        for dept_category in dept_categories_with_unassigned:
            unassigned_count = InvigilationSchedule.objects.filter(
                staff_id__isnull=True,
                hall_dept_category=dept_category
            ).count()
# SAQ
            if unassigned_count > 0:
                assignments_made = reduce_unassigned_slots(dept_category)
                print(
                    f"{dept_category}: Optimization assigned {assignments_made} slots. "
                    f"{unassigned_count - assignments_made} still unassigned."
                )
        double_session()
        categories = ['AIDED', 'SFM', 'SFW']
        for category in categories:
            total_category_slots = InvigilationSchedule.objects.filter(hall_dept_category=category).count()
            assigned_category_slots = InvigilationSchedule.objects.filter(
            hall_dept_category=category, staff_id__isnull=False
            ).count()
            unassigned_category_slots = total_category_slots - assigned_category_slots
        
            print(f"{category}: Assigned {assigned_category_slots}/{total_category_slots} slots.")
            messages.success(request, f"{category}: Assigned {assigned_category_slots}/{total_category_slots} slots.")

        total_slots = InvigilationSchedule.objects.count()
        unassigned_slots = InvigilationSchedule.objects.filter(staff_id__isnull=True).count()
        print(f"Total unassigned slots = {unassigned_slots}/{total_slots}")
        messages.info(request, f"Total unassigned slots = {unassigned_slots}/{total_slots}")
    
    return redirect('generate_schedule')


# Keep these if you still need them for other purposes
def download_assignments(request):
    if 'assignment_log' not in request.session:
        return HttpResponse("No assignment data available")
        
    df = pd.DataFrame(request.session['assignment_log'])
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Assignments', index=False)
    writer.close()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=assignment_log.xlsx'
    return response


def download_room_data(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Room Data"

    headers = ["Hall No", "Department Category", "Department Name", "Strength", "Required Session"]
    ws.append(headers)

    rooms = Room.objects.all().values_list(
        'hall_no', 'dept_category', 'dept_name', 'strength', 'required_session'
    )
    
    for room in rooms:
        ws.append(room)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=RoomData.xlsx'
    wb.save(response)
    return response

def download_staff_unallotted(request):
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Allotment"

    # Headers
    headers = [
        "staff_id", "name", "dept_category", "session",
        "Session 1 Allotted", "Session 2 Allotted",
        "Total Allotted", "Remaining Sessions"
    ]
    ws.append(headers)

    # Loop through all staff
    for staff in Staff.objects.all():
        # Count allotments in invigilation_schedule by staff_id + session number
        session1_allotted = InvigilationSchedule.objects.filter(staff_id=staff.staff_id, session=1).count()
        session2_allotted = InvigilationSchedule.objects.filter(staff_id=staff.staff_id, session=2).count()

        total_allotted = session1_allotted + session2_allotted
        remaining = (staff.session or 0) - total_allotted   # avoid NoneType issue

        ws.append([
            staff.staff_id,
            staff.name,
            staff.dept_category,
            staff.session,
            session1_allotted,
            session2_allotted,
            total_allotted,
            remaining
        ])

    # Return Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="staff_allotment.xlsx"'
    wb.save(response)

    return response

def double_session():
    session_counts = (
        InvigilationSchedule.objects
        .values('date', 'staff_id')
        .annotate(session_count=Count('session', distinct=True))
        .filter(staff_id__isnull=False)
    )

    double_session_staff = {
        (entry['date'], entry['staff_id'])
        for entry in session_counts if entry['session_count'] == 2
    }

    schedules = InvigilationSchedule.objects.all()
    for record in schedules:
        if (record.date, record.staff_id) in double_session_staff:
            record.double_session = True
        else:
            record.double_session = False

    InvigilationSchedule.objects.bulk_update(schedules, ['double_session'])

@require_POST
def swap_slots(request):
    try:
        # Extract selected slot IDs from POST data (expecting exactly two)
        slot_ids = request.POST.getlist('slot_ids[]')  # e.g. ['123', '456']
        if len(slot_ids) != 2:
            return JsonResponse({'success': False, 'message': 'Exactly two slots must be selected for swap.'})
        
        slot1 = InvigilationSchedule.objects.get(pk=slot_ids[0])
        slot2 = InvigilationSchedule.objects.get(pk=slot_ids[1])
        
        # Validate constraint: ensure swapped assignments do not cause staff clash in same session and date
        # Temporarily swap the staff_ids and check for conflicts
        
        # Construct tentative swapped assignments
        swapped_assignments = [
            {'date': slot1.date, 'session': slot1.session, 'staff_id': slot2.staff_id},
            {'date': slot2.date, 'session': slot2.session, 'staff_id': slot1.staff_id},
        ]
        
        for assign in swapped_assignments:
            # Count how many slots this staff would have in same session and date excluding current slots
            conflict_count = InvigilationSchedule.objects.filter(
                date=assign['date'],
                session=assign['session'],
                staff_id=assign['staff_id'],
            ).exclude(pk__in=slot_ids).count()
            
            if conflict_count > 0:
                return JsonResponse({
                    'success': False,
                    'message': f"Swap is not possible for {assign['staff_id']} on {assign['date']} in session {assign['session']}."
                })
        
        # Passed constraints, ask for confirmation on frontend (see JS section below)
        
        
        # Perform swapping inside atomic transaction
        with transaction.atomic():
            temp_staff_id = slot1.staff_id
            temp_name = slot1.name
            temp_designation = slot1.designation
            temp_staff_category = slot1.staff_category
            temp_dept_category = slot1.dept_category
            temp_dept_name = slot1.dept_name
            
            # Swap staff assignments along with related fields
            slot1.staff_id = slot2.staff_id
            slot1.name = slot2.name
            slot1.designation = slot2.designation
            slot1.staff_category = slot2.staff_category
            slot1.dept_category = slot2.dept_category
            slot1.dept_name = slot2.dept_name
            
            slot2.staff_id = temp_staff_id
            slot2.name = temp_name
            slot2.designation = temp_designation
            slot2.staff_category = temp_staff_category
            slot2.dept_category = temp_dept_category
            slot2.dept_name = temp_dept_name
            
            slot1.save()
            slot2.save()
            
            # Update double_session flags after swap
            double_session()
        
        return JsonResponse({'success': True, 'message': 'Slots swapped successfully.'})
    
    except InvigilationSchedule.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'One or both selected slots do not exist.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error during swap: {str(e)}'})